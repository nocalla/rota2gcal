#! python3

# rota2gcal.py
# by Niall O"Callaghan

# gets list of excel files in source folder
# allows choice of file to import
# reads file and creates list of calendar entries for each person in rota file
# gets Google calendar API credentials
# outputs calendar entries for named person from excel file to named calendar

#######################################################

g_weekdays = ["Mon", "Tues", "Wed", "Thurs", "Fri", "Sat", "Sun"]
g_shifts = {
            "9--6": "Early", "10--7": "Middle", "11--8": "Late", 
            "9--7": "Early Long", "10--8": "Late Long", "9--8": "All Day", 
            "9--5": "Early Short", "12--8": "Late Short", "10--6": "Middle Short",
            "off": "Off Work", "off req": "Off Work", "off owed": "Off Work", 
            "training": "Training", "meeting": "Meeting"
            }

import openpyxl
import datetime
import re
import configparser

# Google API stuff
import httplib2
import os

from apiclient import discovery 
from oauth2client import client
from oauth2client import tools
from oauth2client.file import Storage

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/calendar-python-quickstart.json
SCOPES = "https://www.googleapis.com/auth/calendar"
CLIENT_SECRET_FILE = "client_secret.json"
APPLICATION_NAME = "Google Calendar API Python Quickstart"

def get_credentials(): # Google Calendar API Credential stuff - don"t mess with!
    """Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser("~")
    credential_dir = os.path.join(home_dir, ".credentials")
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   "calendar-python-quickstart.json")
    store = Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print("Storing credentials to " + credential_path)
    return credentials

# end of Google API Stuff

def get_configs():
    """ Retrieve all configuration parameters."""
    conf_files = ["rota2gcal.conf"]
    if not os.path.exists("rota2gcal.conf"):
        logging.error("\nError: Can't find configuration file: rota2gcal.conf")
    config = configparser.ConfigParser()
    config.read(conf_files, encoding="utf-8")
    return config

def find_directory(filepath):
    """ finds the downloads folder for the active user if filepath is not set
    """
    if filepath is "":
        if os.name is "nt":
            # Windows
            try:
                import winreg
            except ImportError:
                import _winreg as winreg
            shell_path = ("SOFTWARE\\Microsoft\\Windows\\CurrentVersion"
                          "\\Explorer\\Shell Folders")
            dl_key = "{374DE290-123F-4565-9164-39C4925E467B}"
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, shell_path) as key:
                input_dir = winreg.QueryValueEx(key, dl_key)[0]
        else:
            # Linux, OSX
            userhome = os.path.expanduser('~')
            input_dir = os.path.join(userhome, "Downloads")
    else:
        if not os.path.exists(filepath):
            s = "Error: Input directory not found: {}"
            raise FileNotFoundError(s.format(filepath))
        input_dir = filepath
    return input_dir

def get_rota(dir):
    # gets list of excel files in downloads directory & allows choice
    files = list()
    files_string = "Choose file to import data from by entering relevant number:\n"
    for filename in os.listdir(os.path.normpath(dir)):
        if not filename.endswith(".xlsx"):
            continue
        files.append(filename)
        
    for index, filename in enumerate(files):
        files_string += "{}: {}\n".format(index + 1, filename)
    index = 0
    while True:
        try:
            index = int(input(files_string))       
        except ValueError:
            print("Not an integer!")
            continue
        else:
            index -=1
            break
    f = os.path.join(dir, files[index])
    f = os.path.normpath(f)
    print("Chosen file: {}".format(f))
    return f
    
def parse_rota(f): 
    wb = openpyxl.load_workbook(f)
    sheets = wb.get_sheet_names()
    weeks = list()
    people = dict()

    for sheetname in sheets:
    # strip out extra sheets
        if "Week" in sheetname:
            weeks.append(sheetname)
    
    for sheetname in weeks:
        week = wb.get_sheet_by_name(sheetname)
        date = week.cell(row = 1, column = 2).value
        
        # go through each row, i.e. person
        for row in range(2, week.max_row + 1):
            current_person = week.cell(row = row, column = 1).value
            if current_person:
                if current_person not in people:
                    people[current_person] = list()
                                
                # go through each column, i.e. day
                for column in range(3, week.max_column + 1):
                    day = week.cell(row = 2, column = column).value
                    if day in g_weekdays:
                        val = week.cell(row = row, column = column).value
                        if val and ("=" not in val):
                            day_date = convert_day(date, day, val)
                            if day_date:
                                people[current_person].append(day_date)
    
    # delete people entries with no associated dates 
    people_list = list(people.keys())
    for person in people_list:
        if not people[person]:
            del people[person]
            
    return people
            
def convert_day(wkstart, day, times): 
    # converts day to calendar details
    index = g_weekdays.index(day)
    start_time = float()
    end_time = float()
    output = list()
    all_day = False
    if "--" in times:
        times = times.replace("?", "--")
        times = times.replace(" ", "--")

        time_list = times.split("--")
        times = "{}--{}".format(time_list[0], time_list[1])
        
        try:
            start_time = float(time_list[0])
        except ValueError:
            all_day = True
        # hack here to account for weird formats after the regular time notation
        try:
            end_time = 12 + float(time_list[1])
        except ValueError:
            end_time = 20
    else:
        all_day = True
    
    title = times
    if title in g_shifts:
        # ignore event if it's not covered in our shift templates
        title = g_shifts[title]
        
        start_date = wkstart + datetime.timedelta(days = index, hours = start_time)
        end_date = wkstart + datetime.timedelta(days = index, hours = end_time)
        if all_day:
            start_date = start_date.date().isoformat()
            end_date = end_date.date().isoformat()
        else:
            start_date = start_date.isoformat("T")
            end_date = end_date.isoformat("T")

        output = [start_date, title, end_date, all_day]
    
    return output
    
def get_cal_details(chosen_cal):
    # google calendar api shenanigans!
    # returns [calendar name, id, timezone]
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build("calendar", "v3", http=http)
    
    cal_details = list()
    page_token = None
    while True:
      calendar_list = service.calendarList().list(pageToken=page_token).execute()
      for cal in calendar_list["items"]:
        if cal["summary"] == chosen_cal:
            cal_details = [cal["summary"], cal["id"], cal["timeZone"]]
            break
      page_token = calendar_list.get("nextPageToken")
      if not page_token:
        break
    return cal_details
    
def write_cal_event(event, id):
    # google calendar api shenanigans!
    # returns [calendar name, id, timezone]
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build("calendar", "v3", http=http)
    
    event = service.events().insert(calendarId = id, body = event).execute()
    print("Event created: {}".format(event.get('htmlLink')))
    return
    
class Rota_Calendar(object):
    def __init__(self, config_object):
        self.config = config_object
        self.source_folder = self.get_config_line("Source Folder")
        self.cal_name = self.get_config_line("Calendar Name")
        self.shared_cal_url = self.get_config_line("Shared Calendar URL")
        self.person = self.get_config_line("Person")
        self.event_location = self.get_config_line("Event Location")
        
    def get_config_line(self, key):
        print("{}: {}".format(key, self.config.get("DEFAULT", key))) # debug
        return self.config.get("DEFAULT", key)
        
    def write_to_cal(self, dict, name, t_cal, prefix):
        cal_details = get_cal_details(t_cal)
        if cal_details:
            days = dict[name]        
            # write events to output file & write to calendar
            f = open("last_calendar_output.txt", "w")
            f.write("Calendar details: {}\n".format(cal_details))
            for day in days:
                summary = prefix + day[1]
                startTime = day[0]
                endTime = day[2]
                all_day = day[3]
                timeZone = cal_details[2]
                cal_id = cal_details[1]
                            
                event = {
                  "summary": summary,
                  "start": {
                    "timeZone": timeZone,
                  },
                  "end": {
                    "timeZone": timeZone,
                  },
                }
                # processing for all-day events
                if all_day is True:
                    event["start"]["date"] = startTime
                    event["end"]["date"] = endTime
                else:
                    event["location"] = self.event_location
                    event["start"]["dateTime"] = startTime
                    event["end"]["dateTime"] = endTime
                    
                write_cal_event(event, cal_id)
                f.write("{}\n".format(day))
            f.close()

    def run(self):
        rota_file = get_rota(find_directory(self.source_folder))
        rota_dict = parse_rota(rota_file)
        self.write_to_cal(rota_dict, self.person, self.cal_name, "")
        print("Done.")
    
rota_calendar = Rota_Calendar(get_configs())
rota_calendar.run()
